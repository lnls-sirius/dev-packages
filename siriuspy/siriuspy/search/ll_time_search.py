"""Define properties of all timing devices and their connections."""

import re as _re
from copy import deepcopy as _dcopy

from siriuspy import servweb as _web
from siriuspy.namesys import Filter as _Filter
from siriuspy.namesys import SiriusPVName as _PVName

from siriuspy.search.bpms_search import BPMSearch as _BPMSearch
from siriuspy.search.ps_search import PSSearch as _PSSearch

_timeout = 1.0


class LLTimeSearch:
    """Get the timing devices connections."""

    ll_rgx = _re.compile('([A-Z]+)([0-9]{0,2})', _re.IGNORECASE)

    # defines the relations between input and output of the timing devices
    # that are possible taking into consideration only the devices architecture
    i2o_map = {
        'EVG': {
            'UPLINK': (
                'OUT0', 'OUT1', 'OUT2', 'OUT3',
                'OUT4', 'OUT5', 'OUT6', 'OUT7',
                ),
            },
        'EVR': {
            'UPLINK': (
                'OTP0', 'OTP1', 'OTP2', 'OTP3', 'OTP4', 'OTP5',
                'OTP6', 'OTP7', 'OTP8', 'OTP9', 'OTP10', 'OTP11',
                'OUT0', 'OUT1', 'OUT2', 'OUT3',
                'OUT4', 'OUT5', 'OUT6', 'OUT7',
                ),
            },
        'EVE': {
            'UPLINK': (
                'OUT0', 'OUT1', 'OUT2', 'OUT3',
                'OUT4', 'OUT5', 'OUT6', 'OUT7',
                'RFOUT',
                ),
            },
        'AMCFPGAEVR': {
            'SFP8': (
                'FMC1CH1', 'FMC1CH2', 'FMC1CH3', 'FMC1CH4', 'FMC1CH5',
                'FMC2CH1', 'FMC2CH2', 'FMC2CH3', 'FMC2CH4', 'FMC2CH5',
                'CRT0', 'CRT1', 'CRT2', 'CRT3', 'CRT4',
                'CRT5', 'CRT6', 'CRT7',
                ),
            },
        'OEMultSFP': {
            'OE1': ('OUT1', ),
            'OE2': ('OUT2', ),
            'OE3': ('OUT3', ),
            'OE4': ('OUT4', ),
            },
        'OEMultPOF': {
            'IN1': ('OUT1', ),
            'IN2': ('OUT2', ),
            'IN3': ('OUT3', ),
            'IN4': ('OUT4', ),
            },
        'OESglPOF': {
            'IN': ('OUT', ),
            },
        'OESglSFP': {
            'INRX': ('OUT', 'INTX'),
            },
        'Fout': {
            'UPLINK': (
                'OUT0', 'OUT1', 'OUT2', 'OUT3',
                'OUT4', 'OUT5', 'OUT6', 'OUT7',
                ),
            },
        'PSCtrl': {
            'TIMIN': ('RS485',),
            },
        'Crate': {
            'CRT0': ('CRT0', ),
            'CRT1': ('CRT1', ),
            'CRT2': ('CRT2', ),
            'CRT3': ('CRT3', ),
            'CRT4': ('CRT4', ),
            'CRT5': ('CRT5', ),
            'CRT6': ('CRT6', ),
            'CRT7': ('CRT7', ),
            },
        'OERFTx': {'OpticalACP': ('SIGNAL', )},
        'OERFRx': {'SIGNAL': ('OpticalACP', )},
        }
    i2o_map['FibPatch'] = {
        'P{0:03d}'.format(i): ('P{0:03d}'.format(i), ) for i in range(100)}
    i2o_map['FibPatch']['P052B'] = ('P052B', )

    o2i_map = dict()
    for dev, conns_ in i2o_map.items():
        dic_ = dict()
        o2i_map[dev] = dic_
        for conn1, conns in conns_.items():
            for conn2 in conns:
                dic_[conn2] = conn1
    del(dev, conns_, dic_, conn1, conns, conn2)  # cleanning class namespace.

    _conn_from_evg = None
    _conn_twds_evg = None
    _top_chain_devs = None
    _final_receiver_devs = None
    _all_devices = None
    _hierarchy_map = None

    @classmethod
    def get_channel_input(cls, channel):
        """Get channel input method."""
        if not isinstance(channel, _PVName):
            channel = _PVName(channel)
        o2i = cls.o2i_map.get(channel.dev)
        if o2i is None:
            return []
        conn = o2i.get(channel.propty)
        if conn is None:
            conn = cls.i2o_map[channel.dev].get(channel.propty)
        if conn is None:
            return []
        elif isinstance(conn, str):
            conn = [conn, ]
        return [_PVName(channel.device_name + ':' + co) for co in conn]

    @classmethod
    def add_crates_info(cls, connections_dict=None):
        """Add the information of Crate to BPMs to timing map."""
        cls._get_timedata()
        if connections_dict is None:
            connections_dict = _BPMSearch.get_crates_mapping()
        used = set()
        twds_evg = _dcopy(cls._conn_twds_evg)
        for chan in twds_evg.keys():
            bpms = connections_dict.get(chan.device_name)
            if bpms is None:
                continue
            used.add(chan.device_name)
            for bpm in bpms:
                cls._add_entry_to_map(
                    which_map='from', conn=chan.propty,
                    ele1=chan.device_name, ele2=bpm)
                cls._add_entry_to_map(
                    which_map='twds', conn=chan.propty,
                    ele1=bpm, ele2=chan.device_name)
        cls._update_related_maps()
        return (connections_dict.keys() - used)

    @classmethod
    def add_bbb_info(cls, connections_dict=dict()):
        """Add the information of bbb to PS to timing map."""
        cls._get_timedata()
        if not connections_dict:
            data = _PSSearch.get_bbbname_dict()
            for bbb, bsmps in data.items():
                connections_dict[bbb] = tuple([bsmp[0] for bsmp in bsmps])

        conn = list(cls.i2o_map['PSCtrl'].values())[0][0]
        used = set()
        twds_evg = _dcopy(cls._conn_twds_evg)
        for chan in twds_evg.keys():
            pss = connections_dict.get(chan.device_name)
            if pss is None:
                continue
            used.add(chan.device_name)
            for ps in pss:
                cls._add_entry_to_map(
                    which_map='from', conn=conn,
                    ele1=chan.device_name, ele2=ps)
                cls._add_entry_to_map(
                    which_map='twds', conn=conn,
                    ele1=ps, ele2=chan.device_name)
        cls._update_related_maps()
        return (connections_dict.keys() - used)

    @classmethod
    def get_device_names(cls, filters=None, sorting=None):
        """
        Return a set with all devices of type type_dev.

        if type_dev is None, return all devices.
        """
        cls._get_timedata()
        return _Filter.process_filters(
                            cls._all_devices, filters=filters, sorting=sorting)

    @classmethod
    def get_device_tree(cls, channel):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        twds_evg = cls._conn_twds_evg
        up_chan = _PVName(channel)
        if up_chan in twds_evg.keys():
            up_chan = list(twds_evg[up_chan])[0]
        up_channels = [up_chan]
        while up_chan.device_name not in cls._top_chain_devs:
            up_chan = cls.get_channel_input(up_chan)[0]
            up_chan = list(twds_evg[up_chan])[0]
            up_channels.append(up_chan)
        return up_channels

    @classmethod
    def reset(cls):
        """Reset data to initial value."""
        cls._conn_from_evg = None
        cls._conn_twds_evg = None
        cls._get_timedata()

    @staticmethod
    def server_online():
        """Return True/False if Sirius web server is online."""
        return _web.server_online()

    @classmethod
    def get_connections_from_evg(cls):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        return _dcopy(cls._conn_from_evg)

    @classmethod
    def get_connections_twds_evg(cls):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        return _dcopy(cls._conn_twds_evg)

    @classmethod
    def get_top_chain_senders(cls):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        return _dcopy(cls._top_chain_devs)

    @classmethod
    def get_final_receivers(cls):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        return _dcopy(cls._final_receiver_devs)

    @classmethod
    def get_relations_from_evg(cls):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        return _dcopy(cls._devs_from_evg)

    @classmethod
    def get_relations_twds_evg(cls):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        return _dcopy(cls._devs_twds_evg)

    @classmethod
    def get_hierarchy_list(cls):
        """Return a dictionary with the beaglebone to power supply mapping."""
        cls._get_timedata()
        return _dcopy(cls._hierarchy_map)

    # ############ Auxiliar methods ###########
    @classmethod
    def _add_entry_to_map(cls, which_map, conn, ele1, ele2):
        cls._get_timedata()
        if which_map == 'from':
            mapp = cls._conn_from_evg
        else:
            mapp = cls._conn_twds_evg
        ele1 = _PVName(ele1+':'+conn)
        ele2 = _PVName(ele2+':'+conn)
        ele1_entry = mapp.get(ele1)
        if not ele1_entry:
            mapp[ele1] = {ele2}
        else:
            ele1_entry |= {ele2}

    @classmethod
    def _get_timedata(cls):
        if cls._conn_from_evg:
            return
        if _web.server_online():
            text = _web.timing_devices_mapping(timeout=_timeout)
        cls._parse_text_and_build_connection_mappings(text)
        cls._update_related_maps()

    # Methods auxiliar to _get_timedata
    @classmethod
    def _parse_text_and_build_connection_mappings(cls, text):
        from_evg = dict()
        twds_evg = dict()
        lines = text.splitlines()
        for n, line in enumerate(lines, 1):
            line = line.strip()
            if not line or line[0] == '#':
                continue  # empty line
            out, inn = line.split()[:2]
            out, inn = _PVName(out), _PVName(inn)
            send, ochn, octyp, ocn = cls._get_dev_and_channel(out)
            recv, ichn, ictyp, icn = cls._get_dev_and_channel(inn)
            if not ochn or not ichn:
                print('No {0:s} channel defined in line {1:d}:\n\t {2:s}'
                      .format('output' if not ochn else 'input', n, line))
                return
            elif not octyp or not ictyp:
                print('Sintaxe error in definition of ' +
                      '{0:s} channel in line {1:d}:\n\t {2:s}'
                      .format('output' if not octyp else 'input', n, line))
                return
            else:
                if out in from_evg.keys():
                    from_evg[out] |= {inn}
                else:
                    from_evg[out] = {inn}

                if inn in twds_evg.keys():
                    print('Duplicate device input connection in line ' +
                          '{0:d}:\n\t {1:s}'.format(n, line))
                    return
                else:
                    twds_evg[inn] = {out}
        cls._conn_from_evg = from_evg
        cls._conn_twds_evg = twds_evg

    @classmethod
    def _get_dev_and_channel(cls, txt):
        type_chan = num_chan = None
        dev = txt.device_name
        chan = txt.propty
        reg_match = cls.ll_rgx.findall(chan)
        if reg_match:
            type_chan, num_chan = reg_match[0]
            return dev, chan, type_chan, num_chan
        else:
            print(chan)

    @classmethod
    def _update_related_maps(cls):
        cls._build_devices_relations()
        cls._top_chain_devs = (
            cls._devs_from_evg.keys() - cls._devs_twds_evg.keys())
        cls._final_receiver_devs = (
            cls._devs_twds_evg.keys() - cls._devs_from_evg.keys())
        cls._all_devices = (
            cls._devs_from_evg.keys() | cls._devs_twds_evg.keys())
        cls._build_hierarchy_map()

    @classmethod
    def _build_devices_relations(cls):
        simple_map = dict()
        for k, vs in cls._conn_from_evg.items():
            devs = {v.device_name for v in vs}
            devs |= simple_map.get(k.device_name, set())
            simple_map[k.device_name] = devs

        inv_map = dict()
        for k, vs in simple_map.items():
            for v in vs:
                devs = {k}
                devs |= inv_map.get(v, set())
                inv_map[v] = devs

        cls._devs_from_evg = simple_map
        cls._devs_twds_evg = inv_map

    @classmethod
    def _build_hierarchy_map(cls):
        hierarchy = [cls._top_chain_devs.copy(), ]
        while True:
            vals = set()
            for k in hierarchy[-1]:
                vals |= cls._devs_from_evg.get(k, set())
            if vals:
                hierarchy.append(vals)
            else:
                break
        cls._hierarchy_map = hierarchy


# ###### read excel file. #######
_disclaimer = """
# This file was generated automatically from the data of the
# excel file Cabos_e_Fibras_Sirius.xlsx by the function
# siriuspy.search.ll_time_search.read_excel_file_with_connections.
#
# If the mentioned file change, please, run the script
# again and copy the generated file to replace this one.


"""


def read_excel_file_with_connections():
    from openpyxl import load_workbook
    wb = load_workbook('Cabos_e_Fibras_Sirius.xlsx', data_only=True)
    ws = wb['Cabos e Fibras']

    print(_disclaimer)
    chans = _load_file_and_get_channels(ws)
    chans_sort = _sort_connection_table(chans)
    _print_tables(chans, chans_sort)


def _print_tables(chans, chans_sort):
    print(5*'\n')
    print('# {}'.format(len(chans_sort)))
    for k1, k2 in chans_sort:
        print('{0:35s} {1:35s}'.format(k1, k2))

    print(5*'\n')
    print('# {}'.format(len(chans)))
    for k1, k2 in chans:
        print('# {0:35s} {1:35s}'.format(k1, k2))


def _load_file_and_get_channels(ws):
    chans = list()
    rows = list(ws.iter_rows())
    row0 = rows[0]
    for i, cel in enumerate(row0):
        if cel.value is None:
            continue
        val = cel.value.lower().replace(' ', '')
        if val.endswith('equipamento1'):
            eq1 = i
        elif val.endswith('equipamento2'):
            eq2 = i
        elif val.endswith('porta1'):
            p1 = i
        elif val.endswith('porta2'):
            p2 = i
    for i, row in enumerate(rows[1:]):
        sis = row[0].value
        if sis is None or not sis.lower().startswith(('timing', 'controle')):
            continue
        name1 = _check_device_and_port(row[eq1].value, row[p1].value)
        name2 = _check_device_and_port(row[eq2].value, row[p2].value)
        try:
            name1 = _PVName(name1)
            name2 = _PVName(name2)
        except IndexError:
            print('# {0:04d}:   {1:40s} {2:40s}'.format(i, name1, name2))
            continue
        chans.append((name1, name2))
    return chans


_conversion_linac_names = {
    'LA-RF:H1LLRF': 'LI-RaRF01:RF-LLRFProc',
    'LA-RF:H1SOAM-1': 'LI-RaRF02:RF-SSAmp-1',
    'LA-RF:H1SOAM-2': 'LI-RaRF02:RF-SSAmp-2',
    'LA-RF:H1SOAM-3': 'LI-RaRF02:RF-SSAmp-3',
    'LA-BI:H1FO-1': 'LI-RaDiag02:TI-TrigFout',
    'LA-MD:H1PPS-1': 'LI-RaMD01:MD-PPS',  # ?
    'LA-MD:H1PPS-2': 'LI-RaMD02:MD-PPS',  # ?
    '?': 'IA-00RaCtrl:CO-FibPatch',
    '"Rack" Streak Camera:TI-EVE': 'IA-00RaCtrl:TI-EVE',
    'BO_Dip_05_grupo01:CO-PSCtrl': 'BO-PAD05G01:CO-PSCtrl',
    'BO_Dip_05_grupo02:CO-PSCtrl': 'BO-PAD05G02:CO-PSCtrl',
    'IA-16RaBbB:TI-EVE': 'IA-16RaBbB:TI-EVR',
    }
_translate_port = str.maketrans('', '', ' _-')


def _check_device_and_port(dev, por):
    por = por.upper().translate(_translate_port)
    r_ = _conversion_linac_names.get(dev)
    if r_ is not None:
        dev = r_
    return dev + ':' + por


def _sort_connection_table(chans):
    dev = 'EVG'
    for k1, k2 in chans:
        if k1.dev == dev:
            dev = k1
            break
        elif k2.dev == dev:
            dev = k2
            break
    else:
        raise KeyError('EVG not Found.')

    entries = LLTimeSearch.get_channel_input(
                                        _PVName(dev.device_name+':'+'UPLINK'))
    chans_sort = []
    for entry in entries:
        for i, ks in enumerate(chans):
            k1, k2 = ks
            if k1 == entry:
                entries.extend(LLTimeSearch.get_channel_input(k2))
                chans_sort.append((k1, k2))
                del chans[i]
                break
            if k2 == entry:
                entries.extend(LLTimeSearch.get_channel_input(k1))
                chans_sort.append((k2, k1))
                del chans[i]
                break
    return chans_sort
