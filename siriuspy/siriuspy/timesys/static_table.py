#!/usr/bin/python-sirius

import sys as _sys
import logging as _log

from ..namesys import SiriusPVName as PVName
from ..search import LLTimeSearch


_disclaimer = """
# This file was generated automatically from the data of the
# excel file Cabos_e_Fibras_Sirius.xlsx, located in:
#   https://docs.google.com/spreadsheets/d/19lNNPWxZJv5s-VTrwZRMNWLDMqdHzOQa3ZDIw5neYFI/edit?usp=sharing
#
# The function that created this file can be found in:
#   siriuspy.timesys.static_table.create_static_table.
#
# If the mentioned file change, please, run the script
# again and copy the generated file to replace this one.
"""
_NAMES2CONVERT = {
    'LA-RF:H1TRIG': 'LI-RaRF01:RF-LLRFProc',
    'LA-RF:H1SOAM-1': 'LI-RaRF02:RF-SSAmp-1',
    'LA-RF:H1SOAM-2': 'LI-RaRF02:RF-SSAmp-2',
    'LA-RF:H1SOAM-3': 'LI-RaRF02:RF-SSAmp-3',
    'LA-BI:H1FO-1': 'LI-RaDiag02:TI-TrigFout',
    'LA-MD:H1PPS-1': 'LI-RaMD01:MD-PPS',  # ?
    'LA-MD:H1PPS-2': 'LI-RaMD02:MD-PPS',  # ?
    '?': 'IA-00RaCtrl:CO-DIO',
    '"Rack" Streak Camera:TI-EVE': 'IA-00RaCtrl:TI-EVE'}


def create_static_table(fname=None, local=False, logfile=None):
    """."""
    if logfile:
        hand = _log.FileHandler(logfile, mode='w')
    else:
        hand = _log.StreamHandler(stream=_sys.stdout)
    hand.setFormatter(_log.Formatter('', datefmt=''))
    _log.getLogger().setLevel(level=_log.INFO)
    _log.getLogger().addHandler(hand)

    if local:
        data = read_data_from_local_excel_file(fname)
    else:
        data = read_data_from_google()
    _log.info(_disclaimer)
    chans = _get_channels_from_data(data)
    chans_used, chans_nused = _sort_connection_table(chans)
    _print_tables(chans_used, chans_nused)


def read_data_from_google():
    """Shows basic usage of the Sheets API.
    Prints values from a sample spreadsheet.
    """
    from googleapiclient.discovery import build
    from httplib2 import Http
    from oauth2client import file, client, tools
    _log.getLogger('googleapiclient.discovery_cache').setLevel(_log.ERROR)
    _log.getLogger('googleapiclient.discovery').setLevel(_log.ERROR)
    _log.getLogger('oauth2client.transport').setLevel(_log.ERROR)
    _log.getLogger('oauth2client.client').setLevel(_log.ERROR)

    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    store = file.Storage('/home/fernando/token.json')
    creds = store.get()
    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets(
            '/home/fernando/credentials.json',
            'https://www.googleapis.com/auth/spreadsheets.readonly')
        creds = tools.run_flow(flow, store)
    service = build('sheets', 'v4', http=creds.authorize(Http()))

    # Call the Sheets API
    sheet = service.spreadsheets()
    result = sheet.values().get(
        spreadsheetId='19lNNPWxZJv5s-VTrwZRMNWLDMqdHzOQa3ZDIw5neYFI',
        range='Cabos e Fibras').execute()
    values = result.get('values', [])

    if not values:
        raise ValueError('Error loading file from google')
    return values


def read_data_from_local_excel_file(fname=None):
    from openpyxl import load_workbook
    fname = fname or 'Cabos_e_Fibras_Sirius.xlsx'
    wb = load_workbook(fname, data_only=True)
    ws = wb['Cabos e Fibras']
    data = [[cel.value for cel in row] for row in ws.iter_rows()]
    if not data:
        raise ValueError('Error loading local file')
    return data


def _get_channels_from_data(data):
    for i, cel in enumerate(data[0]):
        if cel is None:
            continue
        val = cel.lower().replace(' ', '')
        if val.endswith('equipamento1'):
            eq1 = i
        elif val.endswith('equipamento2'):
            eq2 = i
        elif val.endswith('porta1'):
            p1 = i
        elif val.endswith('porta2'):
            p2 = i
    chans = list()
    for i, row in enumerate(data[1:]):
        if not row:
            continue
        sis = row[0]
        if sis is None or not sis.lower().startswith(('timing', 'controle')):
            continue
        name1 = _check_device_and_port(row[eq1], row[p1])
        name2 = _check_device_and_port(row[eq2], row[p2])
        try:
            name1 = PVName(name1)
            name2 = PVName(name2)
        except IndexError:
            _log.info('# {0:04d}:   {1:40s} {2:40s}'.format(i, name1, name2))
            continue
        chans.append((name1, name2))
    return chans


def _check_device_and_port(dev, por):
    por = por.upper().translate(str.maketrans('', '', ' _-'))
    return _NAMES2CONVERT.get(dev, dev) + ':' + por


def _sort_connection_table(chans):
    dev = 'EVG'
    for k1, k2 in chans:
        if k1.dev == dev:
            dev = k1
            break
        elif k2.dev == dev:
            dev = k2
            break
    else:
        raise KeyError('EVG not Found.')

    entries = LLTimeSearch.get_channel_input(
        PVName(dev.device_name+':'+'UPLINK'))
    chans_used = []
    for entry in entries:
        mark = list(range(len(chans)))
        for i, ks in enumerate(chans):
            k1, k2 = ks
            if k1 == entry:
                entries.extend(LLTimeSearch.get_channel_input(k2))
                chans_used.append((k1, k2))
                mark.remove(i)
            if k2 == entry:
                entries.extend(LLTimeSearch.get_channel_input(k1))
                chans_used.append((k2, k1))
                mark.remove(i)
        chans = [chans[i] for i in mark]
    chans_nused = sorted(chans)
    return chans_used, chans_nused


def _print_tables(chans_used, chans_nused):
    _log.info(3*'\n')
    _log.info(f'# {len(chans_used):d}')
    for k1, k2 in chans_used:
        _log.info('{0:35s} {1:35s}'.format(k1, k2))

    _log.info(5*'\n')
    _log.info('# CONNECTIONS NOT USED')
    _log.info(f'# {len(chans_nused):d}')
    for k1, k2 in chans_nused:
        _log.info('# {0:35s} {1:35s}'.format(k1, k2))
